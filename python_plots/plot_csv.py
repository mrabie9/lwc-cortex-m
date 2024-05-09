import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
import time
from natsort import natsorted
import openpyxl as xl
from enum import Enum

global x_start_e, x_stop_e, x_start_d, x_stop_d, n_loop, energy_e, energy_d
global i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated, t_calibrations
global n_lcutoff_points, l_cut_off, n_hcutoff_points, h_cut_off, calibrated
global values 

variables = ['n_loop', 'Average encryption time', 'Average decryption time', 'Average encryption energy', 'Average decryption energy',
             'Maximum encryption current', 'Average encryption current', 'Minimum encryption current', 
             'Maximum decryption current', 'Average decryption current', 'Minimum decryption current',
             'Calibrations', 'Calibration time(s)']

# values = [n_loop, t_avg_e, t_avg_d,  energy_e, energy_d,
#           i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated,', '.join(map(str, t_calibrations))]


# Contains column value for the AUT
class AUT(Enum):
    ascon128_enc_dec_1000x = 3
    ascon128Armv7_enc_dec_1000x = 4
    ascon128a_enc_dec_1000x = 5
    ascon128aArmv7_enc_dec_1000x = 6
    isapa128_enc_dec_500x = 7
    isapa128Armv7_enc_dec_500x = 8
    isapa128a_enc_dec_500x = 9
    isapa128aArmv7_enc_dec_500x = 10
    sparkle128_enc_dec_1000x = 11
    sparkle128Armv7_enc_dec_1000x = 12
    sparkle256_enc_dec_1000x = 13
    sparkle256Armv7_enc_dec_1000x = 14
    tinyjambu_enc_dec_1000x = 15
    tinyjambuOpt_enc_dec_1000x = 16
    giftc_enc_dec_200x = 17
    xoodyak_enc_dec_1000x = 18
    romulusn_enc_dec_50x = 19
    romulusnOpt_enc_dec_500x = 20
    eleph_enc_dec_10x = 21
    grain_enc_dec_10x = 22
    photon_enc_dec_15x = 23

# Start col for xlsx write
col = AUT.ascon128_enc_dec_1000x.value 

# Start/end times dict
# data_obtained = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# energy_calc = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O0/ascon128a2_enc_dec_1000x/"
# output = "../Data/Power/python_plots/00_Output/02_run2/O0/ascon128a2_enc_dec_1000x.txt"
# timestamps = {
#     # 'ascon128_enc_dec_1000x': {'start_e': 0.88064, 'stop_e':86.84772 , 'start_d': 89.84808, 'stop_d':174.49783},
#     # 'ascon128a_enc_dec_1000x': {'start_e': 0.52947, 'stop_e':60.11307, 'start_d': 63.11367 , 'stop_d':122.25007},
#     # 'ascon128a2_enc_dec_1000x': {'start_e': 0.71895, 'stop_e': 59.99247, 'start_d':  62.99328, 'stop_d': 122.14987},
#     # 'ascon128a3_enc_dec_1000x': {'start_e': 1.18470, 'stop_e': 59.98670, 'start_d':  62.98678, 'stop_d': 122.63684},
#     # 'giftc_enc_dec_200x': {'start_e': 0.85437, 'stop_e': 156.25025, 'start_d': 159.25074, 'stop_d': 314.85193},
#     'giftc2_enc_dec_200x': {'start_e': 1.25991, 'stop_e': 159.41488, 'start_d': 162.41518, 'stop_d': 320.73839},
#     # 'isapa128_enc_dec_500x': {'start_e': 0.94348, 'stop_e': 141.19371, 'start_d': 144.19465, 'stop_d': 284.03599},
#     # 'isapa128a_enc_dec_500x': {'start_e': 0.99050, 'stop_e': 116.92109, 'start_d':  119.92170, 'stop_d': 235.52180},
#     # 'sparkle128_enc_dec_1000x': {'start_e': 1.22652, 'stop_e': 80.71972, 'start_d': 83.72060, 'stop_d': 163.99806},
#     # 'sparkle256_enc_dec_1000x': {'start_e': 1.04642, 'stop_e': 115.44641, 'start_d': 118.44648, 'stop_d': 223.41102},
#     # 'tinyjambu_enc_dec_1000x': {'start_e': 1.06393, 'stop_e': 136.55366, 'start_d': 139.55458, 'stop_d': 274.98104},
#     # 'xoodyak_enc_dec_1000x': {'start_e': 0.76751, 'stop_e': 217.19979, 'start_d': 220.20037, 'stop_d': 437.68150},
#     # 'eleph_enc_dec_10x': {'start_e': 0.97812, 'stop_e': 120.75210, 'start_d': 123.75230, 'stop_d': 240.38971},
#     # 'grain_enc_dec_10x': {'start_e': 1.06599, 'stop_e': 108.55892, 'start_d': 110.37643, 'stop_d': 211.14548},
#     # 'photon_enc_dec_15x': {'start_e': 1.43538, 'stop_e': 176.73859, 'start_d': 179.73937, 'stop_d': 355.95881},
#     # 'romulusn_enc_dec_50x': {'start_e': 1.32967, 'stop_e': 138.96308, 'start_d': 141.96324, 'stop_d': 279.46629},
    
#     # 'grain2_enc_dec_10x': {'start_e': 1.32234, 'stop_e': 112.35972, 'start_d': 115.36006, 'stop_d': 224.39224},
#     # 'eleph2_enc_dec_10x': {'start_e': 1.18843, 'stop_e': 121.78665, 'start_d': 124.78685, 'stop_d': 245.29173},
#     # 'sparkle2562_enc_dec_1000x': {'start_e': 1.07422, 'stop_e': 115.47440, 'start_d': 118.47448, 'stop_d': 233.43923},

#     # 'eleph3_enc_dec_10x': {'start_e': 7.78885, 'stop_e': 144.36973, 'start_d': 147.37041, 'stop_d': 283.96761},

#     ### 'Random' second sets
#     # 'photon2_enc_dec_15x': {'start_e': 0.78681, 'stop_e': 202.93627, 'start_d': 205.93713, 'stop_d': 407.90221},
#     # 'ascon128a_ngnd_enc_dec_1000x': {'start_e': 0., 'stop_e': , 'start_d': , 'stop_d': }
#     }

data_obtained = True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
energy_calc = True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O0/isapa128a_enc_dec_500x/"
# output = "../Data/Power/python_plots/00_Output/02_run2/O0/isapa128a_enc_dec_500x.txt"
row = 3
timestamps = {
    ## Run 2 start: O0 (with sync) 
    # 'O0': {
    #     'row' : 3,
    #     'ascon128_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'ascon128a_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'isapa128_enc_dec_500x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'isapa128a_enc_dec_500x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'sparkle128_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'sparkle256_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'tinyjambu_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'giftc_enc_dec_200x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'xoodyak_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'romulusn_enc_dec_50x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'eleph_enc_dec_10x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'grain_enc_dec_10x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'photon_enc_dec_15x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    # },

    # 'O2': {
    #     'row' : 107, #20,
    #     'ascon128_enc_dec_1000x': {'start_e': 1.22343, 'stop_e': 12.69106, 'start_d': 15.60643, 'stop_d': 27.87781}, 
    #     'ascon128Armv7_enc_dec_1000x': {'start_e': 1.43493, 'stop_e': 8.52024, 'start_d': 11.52115, 'stop_d': 18.65735}, 
    #     'ascon128a_enc_dec_1000x': {'start_e': 1.50948, 'stop_e': 10.52703, 'start_d': 13.52747, 'stop_d': 22.65489}, 
    #     'ascon128aArmv7_enc_dec_1000x': {'start_e': 1.40604, 'stop_e': 6.02011, 'start_d': 9.02102, 'stop_d': 13.66606}, 
    #     'isapa128_enc_dec_500x': {'start_e': 1.17622, 'stop_e': 49.20612, 'start_d': 52.20624, 'stop_d': 100.22438}, 
    #     'isapa128Armv7_enc_dec_500x': {'start_e': 1.34611, 'stop_e': 13.25041, 'start_d': 16.25112, 'stop_d': 28.19653}, 
    #     'isapa128a_enc_dec_500x': {'start_e': 1.54858, 'stop_e': 43.96354, 'start_d': 46.96358, 'stop_d': 89.36681}, 
    #     'isapa128aArmv7_enc_dec_500x': {'start_e': 1.15943, 'stop_e': 9.99129, 'start_d': 12.99136, 'stop_d': 21.80068}, 
    #     'sparkle128_enc_dec_1000x': {'start_e': 1.40277, 'stop_e': 15.29707, 'start_d': 18.29779, 'stop_d': 32.15008}, 
    #     'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.36627, 'stop_e': 4.45349, 'start_d': 7.45432, 'stop_d': 10.66520}, 
    #     'sparkle256_enc_dec_1000x': {'start_e': 1.29184, 'stop_e': 19.37301, 'start_d': 22.37383, 'stop_d': 41.56206}, 
    #     'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.50430, 'stop_e': 6.07724, 'start_d': 9.07731, 'stop_d': 13.82615}, 
    #     'tinyjambu_enc_dec_1000x': {'start_e': 1.52988, 'stop_e': 30.46323, 'start_d': 33.46392, 'stop_d': 62.39135}, 
    #     'tinyjambuOpt_enc_dec_1000x': {'start_e': 0.97256, 'stop_e': 18.60618, 'start_d': 21.60653, 'stop_d': 39.22403}, 
    #     'giftc_enc_dec_200x': {'start_e': 1.30077, 'stop_e': 69.95490, 'start_d': 72.95574, 'stop_d': 141.77499}, 
    #     'xoodyak_enc_dec_1000x': {'start_e': 1.41285, 'stop_e': 62.05017, 'start_d': 65.05080, 'stop_d': 125.94882}, 
    #     'romulusn_enc_dec_50x': {'start_e': 1.39693, 'stop_e': 26.77422, 'start_d': 29.77494, 'stop_d': 55.19228}, 
    #     'romulusnOpt_enc_dec_500x': {'start_e': 1.48819, 'stop_e': 15.06975, 'start_d': 18.07021, 'stop_d': 31.68071}, 
    #     'eleph_enc_dec_10x': {'start_e': 1.07760, 'stop_e': 41.51174, 'start_d': 44.51259, 'stop_d': 84.94955}, 
    #     'grain_enc_dec_10x': {'start_e': 1.54072, 'stop_e': 33.90685, 'start_d': 36.92842, 'stop_d': 69.27859}, 
    #     'photon_enc_dec_15x': {'start_e': 1.34539, 'stop_e': 81.77962, 'start_d': 84.78039, 'stop_d': 165.22464}, 
    # },

    # 'O3': {
    # #     ## O3 (with sync)
    #     'row' : 108, # 37,
    #     'ascon128_enc_dec_1000x': {'start_e': 1.38242, 'stop_e': 9.31988, 'start_d': 12.32041, 'stop_d': 20.20600}, 
    #     'ascon128Armv7_enc_dec_1000x': {'start_e': 1.36736, 'stop_e':8.53166 , 'start_d': 11.53232, 'stop_d': 18.68690}, 
    #     'ascon128a_enc_dec_1000x': {'start_e': 1.38422, 'stop_e': 6.75916, 'start_d': 9.75925, 'stop_d': 15.23194}, 
    #     'ascon128aArmv7_enc_dec_1000x': {'start_e': 1.38389, 'stop_e': 6.00521, 'start_d': 9.00592, 'stop_d': 13.65451}, 
    #     'isapa128_enc_dec_500x': {'start_e': 1.41481, 'stop_e': 31.31951, 'start_d': 34.31997, 'stop_d': 64.30550}, 
    #     'isapa128Armv7_enc_dec_500x': {'start_e': 1.42632, 'stop_e': 13.04653, 'start_d': 16.04732, 'stop_d': 27.75866}, 
    #     'isapa128a_enc_dec_500x': {'start_e': 1.39807, 'stop_e': 25.75097, 'start_d': 28.75104, 'stop_d': 53.06688}, 
    #     'isapa128aArmv7_enc_dec_500x': {'start_e': 1.43630, 'stop_e': 9.99181, 'start_d': 12.99232, 'stop_d': 21.61812}, 
    #     'sparkle128_enc_dec_1000x': {'start_e': 1.21675, 'stop_e': 12.77109, 'start_d': 15.59983, 'stop_d': 27.80522}, 
    #     'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.02262, 'stop_e': 3.97813, 'start_d': 6.97862, 'stop_d': 10.07170}, 
    #     'sparkle256_enc_dec_1000x': {'start_e': 1.45175, 'stop_e': 18.47239, 'start_d': 21.47278, 'stop_d': 38.90192}, 
    #     'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.45571, 'stop_e': 6.08861, 'start_d': 9.08877, 'stop_d': 13.83345}, 
    #     'tinyjambu_enc_dec_1000x': {'start_e': 1.45466, 'stop_e': 23.19855, 'start_d': 26.19869, 'stop_d': 48.41626}, 
    #     'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.60620, 'stop_e': 18.98973, 'start_d': 21.99025, 'stop_d': 39.57363}, 
    #     'giftc_enc_dec_200x': {'start_e': 1.33235, 'stop_e': 24.81264, 'start_d': 27.81330, 'stop_d': 51.37401}, 
    #     'xoodyak_enc_dec_1000x': {'start_e': 2.24416, 'stop_e': 22.30381, 'start_d': 25.30401, 'stop_d': 44.27070}, # REDO
    #     'romulusn_enc_dec_50x': {'start_e': 1.23536, 'stop_e': 9.91858, 'start_d': 12.91935, 'stop_d': 21.58813}, 
    #     'romulusnOpt_enc_dec_500x': {'start_e': 1.34042, 'stop_e': 14.99041, 'start_d': 17.99042, 'stop_d': 31.62426}, 
    #     'eleph_enc_dec_10x': {'start_e': 1.34734, 'stop_e': 24.68550, 'start_d': 27.68630, 'stop_d': 51.02542}, 
    #     'grain_enc_dec_10x': {'start_e': 1.63439, 'stop_e': 31.80230, 'start_d': 34.80238, 'stop_d': 64.96011}, # REDO
    #     'photon_enc_dec_15x': {'start_e': 1.36881, 'stop_e': 40.00086, 'start_d': 43.00183, 'stop_d': 81.65075}, 
    # }, 

    # 'Os': {
    # #     ## Os (with sync)
    #     'row' : 109, # 54,
    #     'ascon128_enc_dec_1000x': {'start_e': 1.23962, 'stop_e': 19.83298, 'start_d': 22.83364, 'stop_d':41.45027 }, 
    #     'ascon128Armv7_enc_dec_1000x': {'start_e': 0.98255, 'stop_e': 8.19361, 'start_d': 11.19453, 'stop_d': 18.39232}, 
    #     'ascon128a_enc_dec_1000x': {'start_e': 1.22505, 'stop_e': 16.14203, 'start_d': 19.14201, 'stop_d': 34.03150}, 
    #     'ascon128aArmv7_enc_dec_1000x': {'start_e': 1.21922, 'stop_e': 5.87317, 'start_d': 8.87321, 'stop_d': 13.59733}, 
    #     'isapa128_enc_dec_500x': {'start_e': 0.80693, 'stop_e': 61.24708, 'start_d': 64.24792, 'stop_d': 124.37775}, 
    #     'isapa128Armv7_enc_dec_500x': {'start_e': 0.77770, 'stop_e': 12.77343, 'start_d': 15.78580, 'stop_d': 27.86188}, 
    #     'isapa128a_enc_dec_500x': {'start_e': 1.01560, 'stop_e': 53.44280, 'start_d': 56.44359, 'stop_d': 108.48222}, 
    #     'isapa128aArmv7_enc_dec_500x': {'start_e': 1.42883, 'stop_e': 10.31269, 'start_d': 13.31288, 'stop_d': 22.25773}, 
    #     'sparkle128_enc_dec_1000x': {'start_e': 1.33819, 'stop_e': 19.87372, 'start_d': 22.87413, 'stop_d': 42.11158}, 
    #     'sparkle128Armv7_enc_dec_1000x': {'start_e': 0.28369, 'stop_e': 3.52893, 'start_d': 6.52976, 'stop_d': 9.86669}, 
    #     'sparkle256_enc_dec_1000x': {'start_e': 1.51734, 'stop_e': 26.86920, 'start_d': 29.86936, 'stop_d': 55.75462}, 
    #     'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.58264, 'stop_e': 6.24935, 'start_d': 9.24974, 'stop_d': 14.14037}, 
    #     'tinyjambu_enc_dec_1000x': {'start_e': 1.46390, 'stop_e': 26.51556, 'start_d': 29.51591, 'stop_d': 54.79755}, 
    #     'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.44951, 'stop_e': 16.34828, 'start_d': 19.34851, 'stop_d': 34.28753}, 
    #     'giftc_enc_dec_200x': {'start_e': 1.31590, 'stop_e': 58.15518, 'start_d': 61.15587, 'stop_d': 118.10551}, 
    #     'xoodyak_enc_dec_1000x': {'start_e': 1.30736, 'stop_e': 61.90727, 'start_d': 64.90733, 'stop_d': 125.68408}, 
    #     'romulusn_enc_dec_50x': {'start_e': 1.48563, 'stop_e': 40.21259, 'start_d': 43.21261, 'stop_d': 81.98879}, 
    #     'romulusnOpt_enc_dec_500x': {'start_e': 1.37930, 'stop_e': 14.78738, 'start_d': 17.78831, 'stop_d': 31.21841}, 
    #     'eleph_enc_dec_10x': {'start_e': 1.49651, 'stop_e': 58.79010, 'start_d': 61.79051, 'stop_d': 119.08425}, 
    #     'grain_enc_dec_10x': {'start_e': 0.85147, 'stop_e': 37.36903, 'start_d': 40.36945, 'stop_d': 76.56163}, 
    #     'photon_enc_dec_15x': {'start_e': 1.31065, 'stop_e': 78.51346, 'start_d': 81.51364, 'stop_d': 157.61347}, 
    # },
    
    'Os2': {
    #     ## Os (with sync)
        'row' : 74,
        'ascon128_enc_dec_1000x': {'start_e': 2.43930, 'stop_e': 21.03259, 'start_d': 24.03327, 'stop_d': 42.64986}, 
        'ascon128Armv7_enc_dec_1000x': {'start_e': 2.63309, 'stop_e': 9.84420, 'start_d': 12.84509, 'stop_d': 20.04280}, 
        'ascon128a_enc_dec_1000x': {'start_e': 2.47572, 'stop_e': 17.39272, 'start_d': 20.39265, 'stop_d': 35.28217}, 
        'ascon128aArmv7_enc_dec_1000x': {'start_e': 2.58413, 'stop_e': 7.23632, 'start_d': 10.23713, 'stop_d': 14.95947}, 
        'isapa128_enc_dec_500x': {'start_e': 2.62212, 'stop_e': 63.06226, 'start_d':66.06305 , 'stop_d': 126.19294}, 
        'isapa128Armv7_enc_dec_500x': {'start_e': 2.70529, 'stop_e': 14.70106, 'start_d': 17.71256, 'stop_d': 29.81378}, 
        'isapa128a_enc_dec_500x': {'start_e': 2.52588, 'stop_e': 54.95298, 'start_d': 57.95387, 'stop_d': 109.99235}, 
        'isapa128aArmv7_enc_dec_500x': {'start_e': 2.55668, 'stop_e': 11.44059, 'start_d': 14.44068, 'stop_d': 23.38555}, 
        'sparkle128_enc_dec_1000x': {'start_e': 2.48111, 'stop_e': 20.99783, 'start_d': 23.99812, 'stop_d': 43.24300}, 
        'sparkle128Armv7_enc_dec_1000x': {'start_e': 2.58850, 'stop_e': 5.84129, 'start_d': 8.84148, 'stop_d': 12.22471}, 
        'sparkle256_enc_dec_1000x': {'start_e': 2.41419, 'stop_e': 28.00767, 'start_d': 31.00816, 'stop_d': 57.22050}, 
        'sparkle256Armv7_enc_dec_1000x': {'start_e': 2.64796, 'stop_e': 7.30570, 'start_d': 10.30592, 'stop_d': 15.17474}, 
        'tinyjambu_enc_dec_1000x': {'start_e': 2.43576, 'stop_e': 27.48728, 'start_d': 30.48763, 'stop_d': 55.76928}, 
        'tinyjambuOpt_enc_dec_1000x': {'start_e': 2.43026, 'stop_e': 17.40649, 'start_d': 20.40725, 'stop_d': 35.53608}, 
        'giftc_enc_dec_200x': {'start_e': 2.65216, 'stop_e': 59.49165, 'start_d': 62.49212, 'stop_d': 119.44174}, 
        'xoodyak_enc_dec_1000x': {'start_e': 2.01383, 'stop_e': 62.60920, 'start_d': 65.60983, 'stop_d': 126.42357}, 
        'romulusn_enc_dec_50x': {'start_e': 2.59587, 'stop_e': 41.32408, 'start_d': 44.32486, 'stop_d': 83.10189}, 
        'romulusnOpt_enc_dec_500x': {'start_e': 2.59870, 'stop_e': 16.00725, 'start_d': 19.00768, 'stop_d': 32.43824}, 
        'eleph_enc_dec_10x': {'start_e': 2.56502, 'stop_e': 59.71393, 'start_d': 62.71399, 'stop_d': 119.73870}, 
        'grain_enc_dec_10x': {'start_e': 2.52923, 'stop_e': 90.35199, 'start_d': 93.35224, 'stop_d': 180.88921}, 
        'photon_enc_dec_15x': {'start_e': 2.29334, 'stop_e': 79.49743, 'start_d': 82.49835, 'stop_d': 159.71922}, 
    }
}

plot_all_bool = False
Opt = "Os"
run3 = "03_run3"
run2 = "02_run2"
run = run3
# apps = ["ascon128_enc_dec_1000x", "ascon128Armv7_enc_dec_1000x", "ascon128a_enc_dec_1000x", "ascon128aArmv7_enc_dec_1000x"]
# apps = ["isapa128_enc_dec_500x", "isapa128Armv7_enc_dec_500x", "isapa128a_enc_dec_500x", "isapa128aArmv7_enc_dec_500x", ]
# apps = [ "sparkle128_enc_dec_1000x", "sparkle128Armv7_enc_dec_1000x", "sparkle256_enc_dec_1000x", "sparkle256Armv7_enc_dec_1000x"]
# apps = ["tinyjambu_enc_dec_1000x", "tinyjambuOpt_enc_dec_1000x", "giftc_enc_dec_200x", "xoodyak_enc_dec_1000x",]
apps = [ "romulusn_enc_dec_50x" , "romulusnOpt_enc_dec_500x", "eleph_enc_dec_10x","grain_enc_dec_10x", "photon_enc_dec_15x"]

# apps  = ["tinyjambu_enc_dec_1000x"]
def plot_all():
    dfs = []
    t = time.time()
    for app in apps:
        print("Reading ", app, apps.index(app))
        data_dir = "../Data/Power/python_plots/02_csv_dir/"+run+"/" + Opt + "/" + app + "/"
        files = os.listdir(data_dir)

        # Filter out only CSV files
        csv_files = [file for file in files if (file.endswith('.csv') and not "summary" in file)]
        csv_files = natsorted(csv_files)

        
        for file_name in csv_files:
            file_path = os.path.join(data_dir, file_name)
            # Read the CSV file into a DataFrame
            df = pd.read_csv(file_path, delimiter=';', names=[app + "x", app + "y"])
            dfs.append(df)

        # Concatenate all DataFrames into a single DataFrame ignore index HAS to be true
        df = pd.concat(dfs, ignore_index=True)
        # x_values[apps.index(app)] = df['x']/1e3
        # y_values[apps.index(app)] = df['y']/1e6
    print(time.time() - t)
    for app in apps:
        t = time.time()
        x_values= df[app + 'x']/1e3
        y_values = df[app + 'y']/1e6
        # Plotting the graph
        plt.plot(x_values, y_values, linestyle='-') 
        # plt.figure()
        plt.xlabel('Time (s)') 
        plt.ylabel('Current (A)') 
        plt.title(app) 
        # plt.text(4.28073, 0.047, "Start of program", fontsize=12, ha='center', va='center', color='black')
        plt.grid(True)
        print(time.time() - t)
        plt.show()

def main(ts_dict):
    global x_start_e, x_stop_e, x_start_d, x_stop_d, n_loop, energy_e, energy_d
    global i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated, t_calibrations
    global n_lcutoff_points, l_cut_off, n_hcutoff_points, h_cut_off, calibrated
    global values
    # List all files in the directory
    files = os.listdir(data_dir)

    # Filter out only CSV files
    csv_files = [file for file in files if (file.endswith('.csv') and not "summary" in file)]
    csv_files = natsorted(csv_files)
    # print(csv_files)

    # Loop through each CSV file and read its contents
    dfs = []
    for file_name in csv_files:
        file_path = os.path.join(data_dir, file_name)
        # Read the CSV file into a DataFrame
        df = pd.read_csv(file_path, delimiter=';', names=["x", "y"])
        dfs.append(df)

    # Concatenate all DataFrames into a single DataFrame ignore index HAS to be true
    df = pd.concat(dfs, ignore_index=True)
    # Assuming your CSV file has columns named 'x' and 'y', change them accordingly if they are different
    x_values = df['x']/1e3
    y_values = df['y']/1e6
    # print(x_values)

    # Check for calibrations
    n_calibrated = len(y_values[y_values>1])
    t_calibrations = y_values[y_values>1].index
    calibrated = bool(n_calibrated)
    # print("Number of calibrations: ", n_calibrated)

    # remove calibration current samples
    l_cut_off = []
    h_cut_off = 1
    n_hcutoff_points =  len(y_values[y_values>h_cut_off])
    # print("Calibrated: ", calibrated)
    if len(y_values[y_values<0.005]) < 10:
        l_cut_off = 0.005
        n_lcutoff_points =  len(y_values[y_values<l_cut_off])
        x_values = x_values[y_values<1]
        x_values = x_values[y_values>0.005]
        y_values = y_values[y_values<1]
        y_values = y_values[y_values>0.005]
        # print("Removed " + str(n_lcutoff_points) + " at 0.005 A cut-off")
        # print("Removed " + str(n_hcutoff_points) + " above " + str(n_hcutoff_points) + "A")
    else:
        l_cut_off = 0.001
        n_lcutoff_points =  len(y_values[y_values<l_cut_off])
        x_values = x_values[y_values<1]
        x_values = x_values[y_values>0.001]
        y_values = y_values[y_values<1]
        y_values = y_values[y_values>0.001]
        # print("Removed " + str(n_lcutoff_points) + " below " + str(l_cut_off) + "A")
        # print("Removed " + str(n_hcutoff_points) + " above " + str(h_cut_off) + "A")

    ## get start/end time - for report
    # specific_x_val = 3.99519
    # index_of_x = x_values[x_values== specific_x_val].index[0]
    # x_values = x_values[index_of_x:]
    # y_values = y_values[index_of_x:]
    # annotate_x = x_values[index_of_x]
    # annotate_y = round(y_values[index_of_x], 4)
    # # Annotate the point with its coordinates
    # plt.annotate(f'({annotate_x}, {annotate_y})', xy=(annotate_x, annotate_y), xytext=(annotate_x-2e-5, annotate_y+1e-4))
    # plt.xlim(x_values[index_of_x]-9e-5,x_values[index_of_x]+5e-5 )
    # plt.gca().yaxis.set_ticks_position('right')
    # plt.gca().yaxis.set_label_position('right')

    # Figure out n_loop
    if "1000" in data_dir:
        n_loop = 1000
    elif "500" in data_dir:
        n_loop = 500
    elif "200" in data_dir:
        n_loop = 200
    elif "100" in data_dir:
        n_loop = 100
    elif "50" in data_dir:
        n_loop = 50
    elif "15" in data_dir:
        n_loop = 15
    elif "10" in data_dir:
        n_loop = 10
    else: 
        print("N_loop not recognised")
        os._exit

    if energy_calc:
        ## try nested dict
        x_start_e = 0
        x_stop_e = 0
        x_start_d = 0
        x_stop_d = 0

        for app, data in ts_dict.items():
            if app in data_dir:
                x_start_e = data['start_e']
                x_stop_e = data['stop_e']
                x_start_d = data['start_d']
                x_stop_d = data['stop_d']

        # energy calculation - encryption
        # x_start_e = 1.180800
        # x_stop_e = 182.506460
        id_x_start_e = x_values[round(x_values,5)== x_start_e].index[0]
        # print("Start enc time: ", x_values[id_x_start_e])
        id_x_stop_e = x_values[round(x_values,5)== x_stop_e].index[0]
        i_e_avg = round(y_values[id_x_start_e:id_x_stop_e].mean(),5)
        i_e_max = round(y_values[id_x_start_e:id_x_stop_e].max(),5)
        i_e_min = round(y_values[id_x_start_e:id_x_stop_e].min(),5)
        energy_e = i_e_avg * 3.3 * (x_stop_e-x_start_e)/n_loop
        energy_e = round(energy_e,5)
        # print("Average energy - encryption: ", energy_e, (x_stop_e-x_start_e)/n_loop)

        # energy calculation - decryption
        # x_start_d = 185.512820
        # x_stop_d = 367.872190
        id_x_start_d = x_values[round(x_values,5)== x_start_d].index[0]
        id_x_stop_d = x_values[round(x_values,5)== x_stop_d].index[0]
        # print("Start dec time: ", x_values[id_x_start_d])
        i_d_avg = round(y_values[id_x_start_d:id_x_stop_d].mean(),5)
        i_d_max = round(y_values[id_x_start_d:id_x_stop_d].max(),5)
        i_d_min = round(y_values[id_x_start_d:id_x_stop_d].min(),5)
        energy_d = i_d_avg * 3.3 * (x_stop_d-x_start_d)/n_loop
        energy_d = round(energy_d,5)
        # print("Average energy - decryption: ", energy_d, x_start_d, id_x_start_d, x_stop_d, id_x_stop_d, (x_stop_d-x_start_d)/n_loop)

        # # store relevant values
        t_avg_e = round((x_stop_e-x_start_e)/n_loop,5)
        t_avg_d = round((x_stop_d-x_start_d)/n_loop,5)

        # Store results in a text file
        # f = open(output, "w") 
        # f.close()
        # f = open(output, "a")
        # if f.closed:
        #     print("File not open!!")
        # f.write(output)
        # f.write(": \n")
        # f.write("Execution time: \n")
        # f.write("\tEncryption start time:\t\t%f s\n" % x_start_e)
        # f.write("\tEncryption end time:\t\t%f s\n" % x_stop_e)
        # f.write("\tDecryption start time:\t\t%f s\n" % x_start_d)
        # f.write("\tDecryption end time:\t\t%f s\n" % x_stop_d)
        # f.write("\tAverage encryption time:\t%f s\n" % (t_avg_e))
        # f.write("\tAverage decryption time:\t%f s\n\n" % (t_avg_d))
        # f.write("Energy consumption: \n")
        # f.write("\tAverage encryption energy: \t%f J\n" % energy_e)
        # f.write("\tAverage decryption energy: \t%f J\n\n" % energy_d)
        # f.write("Encryption current: \n")
        # f.write("\tMaximum encryption current: %f A\n" % i_e_max)
        # f.write("\tAverage encryption current: %f A\n" % i_e_avg)
        # f.write("\tMinimum encryption current: %f A\n\n" % i_e_min)
        # f.write("Decryption current: \n")
        # f.write("\tMaximum decryption current: %f A\n" % i_d_max)
        # f.write("\tAverage decryption current: %f A\n" % i_d_avg)
        # f.write("\tMinimum decryption current: %f A\n\n" % i_d_min)
        # f.write("Notes: \n")
        # f.write("\tN_LOOP:\t\t\t\t\t\t%d\n" % n_loop)
        # # f.write("\tCalibrated: \t\t\t\t%s\n" % calibrated)
        # f.write("\tNum Calibrations: \t\t\t%d\n" % n_calibrated)
        # f.write("\tCalibration time(s): \t\t")
        # if not calibrated:
        #     f.write("N/A")
        # else:
        #     for t in t_calibrations:
        #         f.write("%.5f s\t" % (round(t/100000,5)))
        #     f.write("\n\tRemoved " + str(n_lcutoff_points) + " samples(s) below " + str(l_cut_off) + "A")
        #     f.write(" and " + str(n_hcutoff_points) + " above " + str(h_cut_off) + "A")
        # f.close
        values = [n_loop, t_avg_e, t_avg_d,  energy_e, energy_d,
          i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated,', '.join(map(str, t_calibrations/100000))]
        # values = [x_start_d - x_stop_e]


    else: 
        # Plotting the graph
        plt.plot(x_values, y_values, linestyle='-')  # You can customize the marker and linestyle as needed
        plt.xlabel('Time (s)')  # Replace 'X Axis Label' with your desired label
        plt.ylabel('Current (A)')  # Replace 'Y Axis Label' with your desired label
        plt.title(data_dir[-15:])  # Replace 'Title of the Graph' with your desired title


        # plt.text(4.28073, 0.047, "Start of program", fontsize=12, ha='center', va='center', color='black')
        plt.grid(True)  # Add gridlines if needed
        plt.show()


def writexl(col, row):
    global values

    # open workbook
    workbook = xl.load_workbook("../Data/Data_m7.xlsx")
    ws = workbook['Power - M7']

    # Write the variable names to the first row
    # for row, var in enumerate(variables, start=row):
    #     ws.cell(row=row, column=col-1, value=var)
    # row = row_start
    for row, var in enumerate(values, start=row):
        ws.cell(row=row, column=col, value=var)
    workbook.save("../Data/Data_m7.xlsx")

if plot_all_bool:
    plot_all()
else:
    if data_obtained:
        t = time.time()
        for opt, data in timestamps.items():
            if "2" in opt:
                run = run3
                opt = opt[:-1]
            else:
                run = run2
            print(opt, run)
            for app, values in data.items():
                if not 'row' in app:
                    # find col
                    col_found = False
                    for col in list(AUT):
                        # print(col.name, app)
                        if app.lower() == col.name.lower():
                            col = col.value
                            col_found = True
                            break
                    if not col_found:
                        if ("romulusnOpt" in app) and ("500x" in app):
                            col = AUT.romulusnOpt_enc_dec_500x.value
                            col_found = True

                    if not col_found:
                        print("Data not stored - col not found!")
                        break
                         
                    print(app)
                    data_dir = "../Data/Power/python_plots/02_csv_dir/" + run + "/" + opt +"/" + app + "/"
                    output = "../Data/Power/python_plots/00_Output/02_run2/" + opt +"/"  + app + ".txt"
                    # print(app)
                    main(data)
                    writexl(col, row)
                    # col +=1
                else:
                    row = values
                    # print("Row ", values)
        print("Took: ", time.time() - t)
    else: 
        main()
